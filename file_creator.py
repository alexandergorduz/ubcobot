import config
import uuid
import urllib3
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU

def create_file(cards, work_cost):
    wb = Workbook()
    ws = wb.active

    cell_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G']

    http = urllib3.PoolManager()

    r = http.request('GET', config.LOGO)
    image_file = io.BytesIO(r.data)
    logo_image = Image(image_file)
    logo_image.width = 220
    logo_image.height = 58
    ws.add_image(logo_image, 'B3')

    font = Font(size=10)
    al_horiz_right = Alignment(horizontal='right')
    al_horiz_vert_center = Alignment(
        horizontal='centerContinuous', 
        vertical='center',
        wrap_text=True
    )
    border = Border(
        left=Side(style='thin', color='00000000'),
        right=Side(style='thin', color='00000000'),
        top=Side(style='thin', color='00000000'),
        bottom=Side(style='thin', color='00000000')
    )
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 36
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 5
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 200
    for num in range(2, 8):
        ws.merge_cells('D{0}:G{0}'.format(num))
    ws['D2'] = 'ТОВ "КОМПАНІЯ УКРБЕЗПЕКА"'
    ws['D2'].font = Font(bold=True)
    ws['D3'] = 'тел. 096 317-77-00, 044 498-57-37'
    ws['D4'] = '099 117-77-93, 093 917-77-66'
    ws['D5'] = '=HYPERLINK("{0}","{1}")'.format('https://ukrbezpeka.kiev.ua', 'ukrbezpeka.kiev.ua')
    ws['D6'] = '=HYPERLINK("{0}","{1}")'.format('https://ukrbezpeka.com', 'ukrbezpeka.com')
    ws['D7'] = '=HYPERLINK("{0}","{1}")'.format('https://t.me/ukrbezpeka', 't.me/ukrbezpeka')
    for num in range(2, 8):
        ws['D{0}'.format(num)].alignment = al_horiz_right
        if num > 4:
            ws['D{0}'.format(num)].font = Font(color='0000FF')
    ws.merge_cells('A10:G10')
    ws['A10'] = 'КОММЕРЧЕСКОЕ ПРЕДЛОЖЕНИЕ'
    ws['A10'].alignment = al_horiz_vert_center
    ws['A10'].font = Font(bold=True, italic=True)
    ws.merge_cells('C12:D12')
    ws['A12'] = '№'
    ws['B12'] = 'Изображение'
    ws['C12'] = 'Наименование'
    ws['E12'] = 'Кол.'
    ws['F12'] = 'Цена'
    ws['G12'] = 'Сумма'
    for letter in cell_letters:
        ws['{0}12'.format(letter)].font = font
        ws['{0}12'.format(letter)].alignment = al_horiz_vert_center
        ws['{0}12'.format(letter)].border = border

    current_row = 12

    for card in cards:
        current_row += 1

        ws.merge_cells('C{0}:D{0}'.format(current_row))
        ws.row_dimensions[current_row].height = 71
        ws['A{0}'.format(current_row)] = card['card_number']
        ws['C{0}'.format(current_row)] = '=HYPERLINK("{0}","{1}")'.format(card['card_url'], card['card_name'])
        ws['E{0}'.format(current_row)] = card['card_amount']
        ws['F{0}'.format(current_row)] = card['card_price']
        ws['G{0}'.format(current_row)] = '=E{0}*F{0}'.format(current_row)
        for letter in cell_letters:
            if letter == 'B' or letter == 'D':
                continue
            else:
                if letter == 'C':
                    ws['{0}{1}'.format(letter, current_row)].font = Font(color='0000FF')
                else:
                    ws['{0}{1}'.format(letter, current_row)].font = font
                ws['{0}{1}'.format(letter, current_row)].alignment = al_horiz_vert_center
                ws['{0}{1}'.format(letter, current_row)].border = border
        ws['B{0}'.format(current_row)].border = border
        ws['D{0}'.format(current_row)].border = border
        r = http.request('GET', card['card_image'])
        image_file = io.BytesIO(r.data)
        image = Image(image_file)
        col = 1
        row = current_row - 1
        col_offset = cm_to_EMU(0.08)
        row_offset = cm_to_EMU(0.2)
        size = XDRPositiveSize2D(pixels_to_EMU(107), pixels_to_EMU(80))
        marker = AnchorMarker(col=col, colOff=col_offset, row=row, rowOff=row_offset)
        image.anchor = OneCellAnchor(_from=marker, ext=size)
        ws.add_image(image)

    for num in range(1, 4):
        ws.row_dimensions[current_row + num].height = 20
        ws.merge_cells('C{0}:D{0}'.format(current_row + num))
        ws.merge_cells('E{0}:G{0}'.format(current_row + num))
    ws.merge_cells('A{0}:A{1}'.format(current_row + 1, current_row + 3))
    ws.merge_cells('B{0}:B{1}'.format(current_row + 1, current_row + 3))
    ws['C{0}'.format(current_row + 1)] = 'Стоимость оборудования, (грн)'
    ws['C{0}'.format(current_row + 2)] = 'Стоимость монтажных работ, (грн)'
    ws['C{0}'.format(current_row + 3)] = 'Общая стоимость, (грн)'
    ws['E{0}'.format(current_row + 1)] = '=SUM(G{0}:G{1})'.format(current_row - len(cards) + 1, current_row)
    ws['E{0}'.format(current_row + 2)] = work_cost
    ws['E{0}'.format(current_row + 3)] = '=SUM(E{0}:G{1})'.format(current_row + 1, current_row + 2)
    for letter in cell_letters:
        for num in range(1, 4):
            ws['{0}{1}'.format(letter, current_row + num)].border = border
            if letter == 'C' or letter == 'E':
                ws['{0}{1}'.format(letter, current_row + num)].font = font
                ws['{0}{1}'.format(letter, current_row + num)].alignment = al_horiz_vert_center
    ws.merge_cells('A{0}:C{0}'.format(current_row + 5))
    ws['A{0}'.format(current_row + 5)] = '*подробнее о товаре Вы можете узнать нажав на наименование'
    ws['A{0}'.format(current_row + 5)].font = Font(size=9, color='FF0000', bold=True, italic=True)

    file_name = 'archive/' + str(uuid.uuid4()) + '.xlsx'
    wb.save(file_name)

    return open(file_name, 'rb')